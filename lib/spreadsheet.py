"""Spreadsheet export for kcommit-analysis-pipeline.

Writes commit data to XLSX, ODS, CSV and related tabular formats.
Column definitions (COMMIT_COLS, COMMIT_COLS_FILTERED, SUMMARY_COLS,
MATRIX_COLS, STATS_COLS) are the canonical source in lib.manifest.

v18.3.0: Report Stats sheet now shows structured context identical to the
         HTML report's top bar and left pane: tool version, run timestamp,
         git range, kernel version, product name, active profiles, score
         threshold, pipeline funnel counts, and coverage metrics.
         Section header rows (bold + light-blue) group related metrics,
         mirroring the HTML sidebar sections.  write_summary_xlsx() and
         write_summary_ods() accept a new optional run_stats kwarg
         (pipeline_run_stats.json content) that provides the extra context;
         callers that omit it fall back to the flat report_stats fields.
"""
import datetime
from lib.scoring import fmt_profiles, fmt_evidence
import os
import zipfile
import xml.sax.saxutils as _sx

# Column definitions imported from manifest (single source of truth)
from lib.manifest import (COMMIT_COLS, COMMIT_COLS_FILTERED,
                          SUMMARY_COLS, MATRIX_COLS, STATS_COLS)


def _profile_scores_text(commit):
    profiles = (((commit or {}).get('scoring') or {}).get('profiles') or {})
    parts = []
    for pname in sorted(profiles):
        try:
            score = float(profiles.get(pname, 0) or 0)
        except (TypeError, ValueError):
            score = 0.0
        parts.append(f"{pname}:{score:g}")
    return '; '.join(parts)


TRACE_COLS = ('sha', 'profile', 'rule', 'matched_level', 'rule_score',
              'profile_score', 'pattern_type', 'pattern', 'matched_value')


# -- Date helpers -------------------------------------------------------------
def _parse_date(ts):
    """Convert a Unix timestamp (int/str) to a datetime object, or None."""
    if not ts:
        return None
    try:
        return datetime.datetime.fromtimestamp(
            int(ts), tz=datetime.timezone.utc).replace(tzinfo=None)
    except (TypeError, ValueError):
        return None


def _fmt_date_str(ts):
    """Return YYYY-MM-DD HH:MM string for display (CSV / ODS text fallback)."""
    dt = _parse_date(ts)
    return dt.strftime("%Y-%m-%d %H:%M") if dt else (str(ts)[:16] if ts else "")


# -- Shared row builders ------------------------------------------------------
def _trace_rows(scored):
    rows = []
    for c in scored or []:
        sha = (c.get('commit') or '')[:12]
        trace = (((c.get('scoring') or {}).get('trace') or {})
                 .get('profiles') or {})
        for pname in sorted(trace):
            pdata = trace.get(pname) or {}
            pscore = pdata.get('final_score', 0)
            rules = pdata.get('rules') or {}
            if not rules:
                rows.append([sha, pname, '', '', 0, pscore, '', '', ''])
                continue
            for rname in sorted(rules):
                rdata = rules.get(rname) or {}
                matches = rdata.get('matches') or {}
                emitted = False
                for kind in ['keywords_whitelist', 'path_whitelist',
                             'commit_whitelist']:
                    for m in (matches.get(kind) or []):
                        rows.append([
                            sha, pname, rname,
                            rdata.get('matched_level', ''),
                            rdata.get('score', 0), pscore,
                            kind, m.get('pattern', ''), m.get('value', ''),
                        ])
                        emitted = True
                if not emitted:
                    rows.append([
                        sha, pname, rname,
                        rdata.get('matched_level', ''),
                        rdata.get('score', 0), pscore, '', '', '',
                    ])
    return rows


def _commit_row(c, include_reason=False, native_types=False):
    """Build a row for a commit record.

    When *native_types* is True (XLSX path), the Date cell is a datetime
    object so openpyxl can apply a real date format.  Otherwise it is a
    formatted string (CSV / ODS text path).
    """
    date_val = (_parse_date(c.get("author_time")) if native_types
                else _fmt_date_str(c.get("author_time")))
    row = [
        c.get("_rank", ""),
        (c.get("commit") or "")[:12],
        c.get("subject", ""),
        c.get("author_name", ""),
        date_val,
        float(c.get("score", 0) or 0),
        fmt_profiles(c),
        _profile_scores_text(c),
        fmt_evidence(c),
    ]
    if include_reason:
        row.append(c.get("_filter_reason", ""))
    return row


def _summary_rows(ps, native_types=False):
    return [
        [n,
         int(d.get("commit_count", d.get("count", 0))),
         float(d.get("total_score", 0)),
         round(float(d.get("avg_score", 0)), 2)]
        for n, d in sorted(ps.items(),
                           key=lambda kv: kv[1].get("commit_count", 0),
                           reverse=True)
    ]


def _matrix_rows(scored, native_types=False):
    rows = []
    for c in scored:
        sc = c.get("scoring", {}) or {}
        for p in (c.get("matched_profiles") or []):
            rows.append([
                c.get("_rank") or "",
                (c.get("commit") or "")[:12],
                c.get("subject", ""),
                p,
                float(c.get("score", 0) or 0),
                float((sc.get("profiles") or {}).get(p, 0)),
            ])
    return rows


# -- Report Stats rows (v18.3.0) ----------------------------------------------
#
# Builds a structured two-column (Metric, Value) table mirroring the HTML
# report top bar and left pane.  Section header rows use the _SECTION sentinel
# and are rendered bold + light-blue in both XLSX and ODS.
#
# report_stats -- dict from st07_report (report_stats.json)
# run_stats    -- optional dict from pipeline_run_stats.json
#                 Provides meta, funnel, stage-level detail.
#                 When None the function falls back to flat report_stats fields.

_SECTION = object()   # sentinel: section header row (bold, no value)


def _stats_rows(report_stats, run_stats=None):
    """Build (metric, value) rows for the Report Stats sheet.

    v18.3.0: structured sections mirror the HTML report context panel.
    Returns a list of (metric, value) tuples; metric may be _SECTION.
    """
    rows = []
    rs  = report_stats or {}
    rns = run_stats    or {}

    def _section(title):
        rows.append((_SECTION, title))

    def _row(metric, value):
        if value is None or value == '' or str(value) == 'None':
            return
        rows.append((metric, value))

    # -- Run Context ----------------------------------------------------------
    _section('Run Context')
    meta = rns.get('meta') or {}
    evl  = rs.get('evaluation') or {}

    _row('Tool version',   meta.get('pipeline_version') or '')
    _row('Generated at',   (meta.get('generated_at') or '')[:19].replace('T', ' '))
    _row('Report title',   rs.get('report_title') or meta.get('product_name') or '')
    _row('Git source',     evl.get('git_source') or '')
    _row('Git range',      evl.get('git_range')  or meta.get('git_range') or '')
    _row('Git baseline',   evl.get('git_baseline') or '')
    _row('Kernel version', meta.get('kernel_version')
                           or evl.get('kernel_revision') or '')

    # -- Analysis Parameters --------------------------------------------------
    _section('Analysis Parameters')
    _row('Active profiles',  evl.get('profiles') or '')
    # Gate threshold/top-n on evl being populated: when evl={} both values
    # would emit spurious hardcoded strings ('0', 'unlimited').
    if evl:
        _row('Score threshold',
             evl.get('min_score') if evl.get('min_score') is not None else '0')
        _row('Top-N limit', str(evl.get('top_n') or 'unlimited'))
    _row('Output formats',   evl.get('outputs') or '')
    _row('HTML detail mode', evl.get('html_detail_mode') or '')

    # -- Pipeline Funnel ------------------------------------------------------
    _section('Pipeline Funnel')
    funnel = rns.get('funnel') or {}
    if funnel:
        _row('Commits collected',        funnel.get('collected'))
        _row('After prefilter',          funnel.get('prefilter_kept'))
        _row('Prefilter dropped',        funnel.get('prefilter_dropped'))
        _row('Scored',                   funnel.get('scored'))
        _row('After postfilter (kept)',  funnel.get('postfilter_kept'))
        _row('Postfilter dropped',       funnel.get('postfilter_dropped'))
        _row('In final report',          funnel.get('final_report'))
        pct = funnel.get('pass_rate_pct')
        _row('Overall pass rate', (str(pct) + '%') if pct is not None else '')
    else:
        # fallback to flat report_stats keys written by st07_report
        _row('Commits collected',  rs.get('st01_collected'))
        _row('After prefilter',    rs.get('st04_prefilter_kept'))
        _row('Prefilter dropped',  rs.get('st04_prefilter_dropped'))
        _row('Scored',             rs.get('st05_total_scored'))
        _row('Postfilter dropped', rs.get('st06_postfilter_dropped'))
        _row('In final report',    rs.get('total_scored_commits'))

    # -- Scoring Summary ------------------------------------------------------
    _section('Scoring Summary')
    st05 = rns.get('stage_05_scoring') or {}
    if st05:
        _row('Score max',             st05.get('score_max'))
        _row('Score min',             st05.get('score_min'))
        _row('Score avg',             st05.get('score_avg'))
        _row('Score median',          st05.get('score_median'))
        _row('Zero-score commits',    st05.get('zero_score_commits'))
        _row('Multi-profile commits', st05.get('multi_profile_commits'))
    else:
        _row('Score highest', rs.get('score_highest'))
        _row('Score lowest',  rs.get('score_lowest'))
        _row('Score avg',     rs.get('score_avg'))

    # -- Coverage -------------------------------------------------------------
    _section('Coverage')
    _row('Commits with no profile match',
         rs.get('commits_matched_zero_profiles'))
    _row('Commits with product evidence',
         rs.get('commits_with_product_evidence'))
    pm = rns.get('product_map_summary') or {}
    if pm:
        _row('KConfig symbols enabled', pm.get('kconfig_symbols_enabled'))
        _row('Compiled files tracked',  pm.get('compiled_files'))
        _row('Compiled dirs tracked',   pm.get('compiled_dirs'))

    return rows


# -- XLSX via openpyxl --------------------------------------------------------
_XLSX_DATE_FMT  = "YYYY-MM-DD HH:MM"
_XLSX_FLOAT_FMT = "0.00"
_XLSX_INT_FMT   = "0"


def _xlsx_write_sheet(ws, headers, rows):
    """Write headers + rows to *ws* with auto-width and typed cell formats."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT_W = Font(bold=True, name="Calibri", size=11, color="FFFFFF")
    TOP           = Alignment(vertical="top", wrap_text=False)

    ws.append(headers)
    for cell in ws[1]:
        cell.font      = HEADER_FONT_W
        cell.fill      = HEADER_FILL
        cell.alignment = TOP
    ws.row_dimensions[1].height = 18

    for row in rows:
        ws.append(row)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = TOP
            if isinstance(cell.value, datetime.datetime):
                cell.number_format = _XLSX_DATE_FMT
            elif isinstance(cell.value, float):
                cell.number_format = _XLSX_FLOAT_FMT
            elif isinstance(cell.value, int) and not isinstance(cell.value, bool):
                cell.number_format = _XLSX_INT_FMT

    for i, col_cells in enumerate(ws.columns):
        col_letter = get_column_letter(i + 1)
        max_len = 0
        for cell in col_cells:
            if cell.value is None:
                continue
            display = (cell.value.strftime("%Y-%m-%d %H:%M")
                       if isinstance(cell.value, datetime.datetime)
                       else str(cell.value))
            max_len = max(max_len, len(display))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 60)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _xlsx_write_stats_sheet(ws, rows):
    """Write the Report Stats sheet with section header rows styled distinctly.

    rows -- output of _stats_rows(): list of (metric, value) tuples
            where metric may be the _SECTION sentinel.
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    HEADER_FILL  = PatternFill('solid', fgColor='1F4E79')
    HEADER_FONT  = Font(bold=True, name='Calibri', size=11, color='FFFFFF')
    SECTION_FILL = PatternFill('solid', fgColor='D6E4F0')
    SECTION_FONT = Font(bold=True, name='Calibri', size=10, color='1F4E79')
    TOP          = Alignment(vertical='top', wrap_text=False)

    ws.append(['Metric', 'Value'])
    for cell in ws[1]:
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = TOP
    ws.row_dimensions[1].height = 18

    for metric, value in rows:
        if metric is _SECTION:
            ws.append([value, ''])
            r = ws.max_row
            for cell in ws[r]:
                cell.font      = SECTION_FONT
                cell.fill      = SECTION_FILL
                cell.alignment = TOP
        else:
            ws.append([metric, value])
            r = ws.max_row
            ws.cell(r, 1).alignment = TOP
            vc = ws.cell(r, 2)
            vc.alignment = TOP
            if isinstance(value, float):
                vc.number_format = _XLSX_FLOAT_FMT
            elif isinstance(value, int) and not isinstance(value, bool):
                vc.number_format = _XLSX_INT_FMT

    for i, col_cells in enumerate(ws.columns):
        col_letter = get_column_letter(i + 1)
        max_len = max(
            (len(str(c.value)) for c in col_cells if c.value is not None),
            default=8,
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 80)

    ws.freeze_panes = 'A2'


def _xlsx_save(wb, path):
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    wb.save(path)


def _new_wb():
    import openpyxl
    return openpyxl.Workbook()


def write_xlsx(path: str, scored: list, profile_summary: dict,
               sheet_name: str = "Commits",
               include_reason: bool = False) -> None:
    """Write a single-sheet XLSX for *scored* commits."""
    wb = _new_wb()
    ws = wb.active
    ws.title = sheet_name
    cols = COMMIT_COLS_FILTERED if include_reason else COMMIT_COLS
    _xlsx_write_sheet(ws, cols,
                      [_commit_row(c, include_reason, native_types=True)
                       for c in scored])
    _xlsx_save(wb, path)


def write_profile_summary_xlsx(path: str, profile_summary: dict) -> None:
    """Write a single-sheet XLSX for profile summary."""
    wb = _new_wb()
    ws = wb.active
    ws.title = "Profile Summary"
    _xlsx_write_sheet(ws, SUMMARY_COLS,
                      _summary_rows(profile_summary, native_types=True))
    _xlsx_save(wb, path)


def write_profile_matrix_xlsx(path: str, scored: list) -> None:
    """Write a single-sheet XLSX for profile matrix."""
    wb = _new_wb()
    ws = wb.active
    ws.title = "Profile Matrix"
    _xlsx_write_sheet(ws, MATRIX_COLS, _matrix_rows(scored, native_types=True))
    _xlsx_save(wb, path)


def write_summary_xlsx(path: str, scored: list, filtered: list,
                       profile_summary: dict, report_stats: dict = None,
                       report_title: str = "kcommit Analysis Report",
                       run_stats: dict = None) -> None:
    """Write a multi-sheet summary XLSX.

    Sheet order: Report Stats, Relevant Commits, Filtered Commits,
                 Profile Summary, Profile Matrix, Rule Trace.

    run_stats -- optional pipeline_run_stats.json content (v18.3.0);
                 when supplied the Report Stats sheet shows full context
                 mirroring the HTML report top bar and left pane.
    """
    wb = _new_wb()

    ws0 = wb.active
    ws0.title = "Report Stats"
    _xlsx_write_stats_sheet(ws0, _stats_rows(report_stats, run_stats=run_stats))

    ws1 = wb.create_sheet("Relevant Commits")
    _xlsx_write_sheet(ws1, COMMIT_COLS,
                      [_commit_row(c, native_types=True) for c in scored])

    if filtered:
        ws2 = wb.create_sheet("Filtered Commits")
        _xlsx_write_sheet(ws2, COMMIT_COLS_FILTERED,
                          [_commit_row(c, include_reason=True, native_types=True)
                           for c in filtered])

    if profile_summary:
        ws3 = wb.create_sheet("Profile Summary")
        _xlsx_write_sheet(ws3, SUMMARY_COLS,
                          _summary_rows(profile_summary, native_types=True))

    ws4 = wb.create_sheet("Profile Matrix")
    _xlsx_write_sheet(ws4, MATRIX_COLS, _matrix_rows(scored, native_types=True))

    ws5 = wb.create_sheet("Rule Trace")
    _xlsx_write_sheet(ws5, TRACE_COLS, _trace_rows(scored))

    _xlsx_save(wb, path)


# -- ODS via stdlib zipfile ---------------------------------------------------
_ODS_NS = (
    ' xmlns:office="urn:oasis:names:tc:opendocument:xmlns:office:1.0"'
    ' xmlns:table="urn:oasis:names:tc:opendocument:xmlns:table:1.0"'
    ' xmlns:text="urn:oasis:names:tc:opendocument:xmlns:text:1.0"'
    ' xmlns:fo="urn:oasis:names:tc:opendocument:xmlns:xsl-fo-compatible:1.0"'
    ' xmlns:style="urn:oasis:names:tc:opendocument:xmlns:style:1.0"'
    ' xmlns:number="urn:oasis:names:tc:opendocument:xmlns:datastyle:1.0"'
)
_ODS_HEAD = (
    '<?xml version="1.0" encoding="UTF-8"?>'
    f'<office:document-content{_ODS_NS} office:version="1.2">'
    '<office:automatic-styles>'
    # Bold column header cell
    '<style:style style:name="H" style:family="table-cell">'
    '<style:text-properties fo:font-weight="bold"/></style:style>'
    # Section header cell: bold blue text + light-blue background
    '<style:style style:name="SH" style:family="table-cell">'
    '<style:table-cell-properties fo:background-color="#D6E4F0"/>'
    '<style:text-properties fo:font-weight="bold" fo:color="#1F4E79"/>'
    '</style:style>'
    # Auto-width column
    '<style:style style:name="CO" style:family="table-column">'
    '<style:table-column-properties'
    ' style:use-optimal-column-width="true"/></style:style>'
    # Date format
    '<number:date-style style:name="ND">'
    '<number:year number:style="long"/>'
    '<number:text>-</number:text>'
    '<number:month number:style="long"/>'
    '<number:text>-</number:text>'
    '<number:day number:style="long"/>'
    '<number:text> </number:text>'
    '<number:hours number:style="long"/>'
    '<number:text>:</number:text>'
    '<number:minutes number:style="long"/>'
    '</number:date-style>'
    '<style:style style:name="DC" style:family="table-cell"'
    ' style:data-style-name="ND"/>'
    # Float format
    '<number:number-style style:name="NF">'
    '<number:number number:decimal-places="2" number:min-integer-digits="1"/>'
    '</number:number-style>'
    '<style:style style:name="FC" style:family="table-cell"'
    ' style:data-style-name="NF"/>'
    '</office:automatic-styles>'
    '<office:body><office:spreadsheet>'
)
_ODS_TAIL = '</office:spreadsheet></office:body></office:document-content>'


def _ods_cell(value, bold=False, section=False):
    """Render a typed ODS cell element.

    section=True applies the SH style (bold blue text + light-blue fill).
    bold=True applies the H style (bold text only, for column headers).
    """
    if section:
        style_attr = ' table:style-name="SH"'
    elif bold:
        style_attr = ' table:style-name="H"'
    else:
        style_attr = ''

    if isinstance(value, datetime.datetime):
        iso  = value.strftime("%Y-%m-%dT%H:%M:%S")
        disp = _sx.escape(value.strftime("%Y-%m-%d %H:%M"))
        return (f'<table:table-cell table:style-name="DC"'
                f' office:value-type="date" office:date-value="{iso}">'
                f'<text:p>{disp}</text:p></table:table-cell>')
    if isinstance(value, float):
        esc = _sx.escape(f"{value:.2f}")
        # float cells: prefer explicit FC style; section/bold override not
        # meaningful for numeric cells so we apply FC regardless.
        return (f'<table:table-cell table:style-name="FC"'
                f' office:value-type="float" office:value="{value}">'
                f'<text:p>{esc}</text:p></table:table-cell>')
    if isinstance(value, int) and not isinstance(value, bool):
        esc = _sx.escape(str(value))
        return (f'<table:table-cell{style_attr}'
                f' office:value-type="float" office:value="{value}">'
                f'<text:p>{esc}</text:p></table:table-cell>')
    esc = _sx.escape(str(value) if value is not None else "")
    return (f'<table:table-cell{style_attr} office:value-type="string">'
            f'<text:p>{esc}</text:p></table:table-cell>')


def _ods_sheet(name, headers, rows):
    ncols    = len(headers)
    col_tags = '<table:table-column table:style-name="CO"/>' * ncols
    lines = [
        f'<table:table table:name="{_sx.escape(name)}\">'
        + col_tags
        + '<table:table-row>'
        + ''.join(_ods_cell(h, bold=True) for h in headers)
        + '</table:table-row>'
    ]
    for row in rows:
        lines.append(
            '<table:table-row>'
            + ''.join(_ods_cell(v) for v in row)
            + '</table:table-row>'
        )
    lines.append('</table:table>')
    return ''.join(lines)


def _ods_stats_sheet(rows):
    """Render the Report Stats sheet as ODS XML with styled section headers.

    rows -- output of _stats_rows(): list of (metric, value) tuples
            where metric may be the _SECTION sentinel.
    """
    col_tags = '<table:table-column table:style-name="CO"/>' * 2
    lines = [
        '<table:table table:name="Report Stats">'
        + col_tags
        + '<table:table-row>'
        + _ods_cell('Metric', bold=True)
        + _ods_cell('Value',  bold=True)
        + '</table:table-row>'
    ]
    for metric, value in rows:
        if metric is _SECTION:
            lines.append(
                '<table:table-row>'
                + _ods_cell(value, section=True)
                + _ods_cell('',    section=True)
                + '</table:table-row>'
            )
        else:
            lines.append(
                '<table:table-row>'
                + _ods_cell(metric)
                + _ods_cell(value)
                + '</table:table-row>'
            )
    lines.append('</table:table>')
    return ''.join(lines)


def _ods_save(content, path):
    manifest = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<manifest:manifest'
        ' xmlns:manifest="urn:oasis:names:tc:opendocument:xmlns:manifest:1.0">'
        '<manifest:file-entry'
        ' manifest:media-type="application/vnd.oasis.opendocument.spreadsheet"'
        ' manifest:full-path="/"/>'
        '<manifest:file-entry manifest:media-type="text/xml"'
        ' manifest:full-path="content.xml"/>'
        '</manifest:manifest>'
    )
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zi = zipfile.ZipInfo("mimetype")
        zi.compress_type = zipfile.ZIP_STORED
        zf.writestr(zi, "application/vnd.oasis.opendocument.spreadsheet")
        zf.writestr("META-INF/manifest.xml", manifest)
        zf.writestr("content.xml", content)


def write_ods(path: str, scored: list, profile_summary: dict,
              sheet_name: str = "Commits",
              include_reason: bool = False) -> None:
    """Write a single-sheet ODS for *scored* commits."""
    cols = COMMIT_COLS_FILTERED if include_reason else COMMIT_COLS
    content = (
        _ODS_HEAD
        + _ods_sheet(sheet_name, cols,
                     [_commit_row(c, include_reason, native_types=True)
                      for c in scored])
        + _ODS_TAIL
    )
    _ods_save(content, path)


def write_profile_summary_ods(path: str, profile_summary: dict) -> None:
    """Write a single-sheet ODS for profile summary."""
    content = (
        _ODS_HEAD
        + _ods_sheet("Profile Summary", SUMMARY_COLS,
                     _summary_rows(profile_summary, native_types=True))
        + _ODS_TAIL
    )
    _ods_save(content, path)


def write_profile_matrix_ods(path: str, scored: list) -> None:
    """Write a single-sheet ODS for profile matrix."""
    content = (
        _ODS_HEAD
        + _ods_sheet("Profile Matrix", MATRIX_COLS,
                     _matrix_rows(scored, native_types=True))
        + _ODS_TAIL
    )
    _ods_save(content, path)


def write_summary_ods(path: str, scored: list, filtered: list,
                      profile_summary: dict, report_stats: dict = None,
                      report_title: str = "kcommit Analysis Report",
                      run_stats: dict = None) -> None:
    """Write a multi-sheet summary ODS.

    Sheet order: Report Stats, Relevant Commits, Filtered Commits,
                 Profile Summary, Profile Matrix.

    run_stats -- optional pipeline_run_stats.json content (v18.3.0);
                 when supplied the Report Stats sheet shows full context
                 mirroring the HTML report top bar and left pane.
    """
    sheets  = _ods_stats_sheet(_stats_rows(report_stats, run_stats=run_stats))
    sheets += _ods_sheet("Relevant Commits", COMMIT_COLS,
                         [_commit_row(c, native_types=True) for c in scored])
    if filtered:
        sheets += _ods_sheet("Filtered Commits", COMMIT_COLS_FILTERED,
                             [_commit_row(c, include_reason=True,
                                          native_types=True)
                              for c in filtered])
    if profile_summary:
        sheets += _ods_sheet("Profile Summary", SUMMARY_COLS,
                             _summary_rows(profile_summary, native_types=True))
    sheets += _ods_sheet("Profile Matrix", MATRIX_COLS,
                         _matrix_rows(scored, native_types=True))
    _ods_save(_ODS_HEAD + sheets + _ODS_TAIL, path)
